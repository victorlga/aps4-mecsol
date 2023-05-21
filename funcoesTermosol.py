import numpy as np
# -*- coding: utf-8 -*-
"""
A funcao 'plota' produz um gráfico da estrutura definida pela matriz de nos N 
e pela incidencia Inc.

Sugestao de uso:

from funcoesTermosol import plota
plota(N,Inc)
-------------------------------------------------------------------------------
A funcao 'importa' retorna o numero de nos [nn], a matriz dos nos [N], o numero
de membros [nm], a matriz de incidencia [Inc], o numero de cargas [nc], o vetor
carregamento [F], o numero de restricoes [nr] e o vetor de restricoes [R] 
contidos no arquivo de entrada.

Sugestao de uso:
    
from funcoesTermosol import importa
[nn,N,nm,Inc,nc,F,nr,R] = importa('entrada.xlsx')
-------------------------------------------------------------------------------
A funcao 'geraSaida' cria um arquivo nome.txt contendo as reacoes de apoio Ft, 
deslocamentos Ut, deformacoes Epsi, forcas Fi e tensoes Ti internas. 
As entradas devem ser vetores coluna.

Sugestao de uso:
    
from funcoesTermosol import geraSaida
geraSaida(nome,Ft,Ut,Epsi,Fi,Ti)
-------------------------------------------------------------------------------

"""
def plota(N,Inc):
    # Numero de membros
    nm = len(Inc[:,0])
    
    import matplotlib as mpl
    import matplotlib.pyplot as plt

#    plt.show()
    fig = plt.figure()
    # Passa por todos os membros
    for i in range(nm):
        
        # encontra no inicial [n1] e final [n2] 
        n1 = int(Inc[i,0])
        n2 = int(Inc[i,1])        

        plt.plot([N[0,n1-1],N[0,n2-1]],[N[1,n1-1],N[1,n2-1]],color='r',linewidth=3)

    plt.xlabel('x [m]')
    plt.ylabel('y [m]')
    plt.grid(True)
    plt.axis('equal')
    plt.show()

def gauss_seidel(A, b, x0, epsilon, max_iterations):
    n = len(A)
    x = x0.copy()

    #Gauss-Seidal Method [By Bottom Science]

    for i in range(max_iterations):
        x_new = np.zeros(n)
        for j in range(n):
            s1 = np.dot(A[j, :j], x_new[:j])
            s2 = np.dot(A[j, j + 1:], x[j + 1:])
            x_new[j] = (b[j] - s1 - s2) / A[j, j]
        if np.allclose(x, x_new, rtol=epsilon):
            return x_new
        x = x_new
    return x


def convert_xlsx_to_xls(input_file, output_file):
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows    # Load the Excel file
    wb = openpyxl.load_workbook(input_file)
    
    # Create a new Workbook for saving in .xls format
    xls_wb = openpyxl.Workbook()
    
    # Copy each sheet to the new Workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        xls_sheet = xls_wb.create_sheet(title=sheet_name)
        
        for row in sheet.iter_rows(values_only=True):
            xls_sheet.append(row)
    
    # Remove the default sheet created by openpyxl
    del xls_wb['Sheet']
    
    # Save the new Workbook as .xls file
    xls_wb.save(output_file)



def importa(entradaNome):
    import numpy as np
    import pandas as pd
    # Read the Excel file
    
    ################################################## Ler os nos
    nos = pd.read_excel(entradaNome, sheet_name='Nos')
    
    
    # Numero de nos
    nn = int(nos.iat[0, 3])
                 
    # Matriz dos nós
    N = np.zeros((2, nn))
    
    for c in range(nn):
        N[0, c] = nos.iat[c, 0]
        N[1, c] = nos.iat[c, 1]
    
    ################################################## Ler a incidencia
    incid = pd.read_excel(entradaNome, sheet_name='Incidencia')
    
    # Numero de membros
    nm = int(incid.iat[0, 5])
                 
    # Matriz de incidencia
    Inc = np.zeros((nm, 4))
    
    for c in range(nm):
        Inc[c, 0] = int(incid.iat[c, 0])
        Inc[c, 1] = int(incid.iat[c, 1])
        Inc[c, 2] = incid.iat[c, 2]
        Inc[c, 3] = incid.iat[c, 3]
    
    ################################################## Ler as cargas
    carg = pd.read_excel(entradaNome, sheet_name='Carregamento')
    
    # Numero de cargas
    nc = int(carg.iat[0, 4])
                 
    # Vetor carregamento
    F = np.zeros((nn*2, 1))
    
    for c in range(nc):
        no = carg.iat[c, 0]
        xouy = carg.iat[c, 1]
        GDL = int(no*2 - (2 - xouy))
        F[GDL-1, 0] = carg.iat[c, 2]
         
    ################################################## Ler restricoes
    restr = pd.read_excel(entradaNome, sheet_name='Restricao')
    
    # Numero de restricoes
    nr = int(restr.iat[0, 3])
                 
    # Vetor com os graus de liberdade restritos
    R = np.zeros((nn*2, 1))
    
    for c in range(nr):
        no = restr.iat[c, 0]
        xouy = restr.iat[c, 1]
        GDL = int(no*2 - (2 - xouy))
        R[GDL-1, 0] = 1
        

    return nn, N, nm, Inc, nc, F, nr, R

def geraSaida(nome,Ft,Ut,Epsi,Fi,Ti):
    nome = nome + '.txt'
    f = open("saida.txt","w+")
    f.write('Reacoes de apoio [N]\n')
    f.write(str(Ft))
    f.write('\n\nDeslocamentos [m]\n')
    f.write(str(Ut))
    f.write('\n\nDeformacoes []\n')
    f.write(str(Epsi))
    f.write('\n\nForcas internas [N]\n')
    f.write(str(Fi))
    f.write('\n\nTensoes internas [Pa]\n')
    f.write(str(Ti))
    f.close()
    

def solve_system(N, Inc, nc, F, nr, R):
    # Number of nodes
    nn = len(N[0])

    # Number of members
    nm = len(Inc)

    # Create the global stiffness matrix
    K = np.zeros((2*nn, 2*nn))

    for i in range(nm):
        n1 = int(Inc[i, 0])
        n2 = int(Inc[i, 1])
        E = Inc[i, 2]
        A = Inc[i, 3]

        x1 = N[0, n1-1]
        y1 = N[1, n1-1]
        x2 = N[0, n2-1]
        y2 = N[1, n2-1]

        L = np.sqrt((x2 - x1)**2 + (y2 - y1)**2)
        c = (x2 - x1) / L
        s = (y2 - y1) / L

        k = (E * A / L) * np.array([
            [c**2, c*s, -c**2, -c*s],
            [c*s, s**2, -c*s, -s**2],
            [-c**2, -c*s, c**2, c*s],
            [-c*s, -s**2, c*s, s**2]
        ])

        GDL = np.array([
            [2*n1 - 1],
            [2*n1],
            [2*n2 - 1],
            [2*n2]
        ])

        K[np.ix_(GDL.flatten(), GDL.flatten())] += k

    # Separate the prescribed and unknown displacements
    R = R.flatten().astype(int)
    U_known = np.zeros((nr,))
    U_unknown = np.zeros((2*nn - nr,))

    U_known = F[R]
    U_unknown_idx = np.setdiff1d(np.arange(2*nn), R)
    U_unknown = np.copy(F[U_unknown_idx])

    # Partition the stiffness matrix
    K11 = K[np.ix_(U_unknown_idx, U_unknown_idx)]
    K12 = K[np.ix_(U_unknown_idx, R)]
    K21 = K[np.ix_(R, U_unknown_idx)]
    K22 = K[np.ix_(R, R)]

    # Solve the system of equations using the Jacobi method
    tol = 1e-6
    max_iterations = 1000

    D = np.diag(K11)
    LU = K11 - np.diag(D)
    U_unknown_new = np.copy(U_unknown)
    U_unknown_prev = np.copy(U_unknown)

    for iteration in range(max_iterations):
        U_unknown_new = (U_known - K21 @ U_known - LU @ U_unknown_prev) / D

        if np.linalg.norm(U_unknown_new - U_unknown_prev) < tol:
            break

        U_unknown_prev = np.copy(U_unknown_new)

    U_unknown = U_unknown_new

    # Assemble the complete displacement vector
    U = np.zeros((2*nn,))
    U[U_unknown_idx] = U_unknown
    U[R] = U_known

    # Compute reactions at supports
    Ft = K22 @ U_known + K21 @ U_unknown

    return Ft, U


if __name__ == "__main__":
    [nn, N, nm, Inc, nc, F, nr, R] = importa('entrada.xlsx')
    Ft, Ut = solve_system(N, Inc, nc, F, nr, R)
    Epsi = np.zeros((nm,))
    Fi = np.zeros((nm,))
    Ti = np.zeros((nm,))

    for i in range(nm):
        n1 = int(Inc[i, 0])
        n2 = int(Inc[i, 1])
        E = Inc[i, 2]
        A = Inc[i, 3]
        x1 = N[0, n1-1]
        y1 = N[1, n1-1]
        x2 = N[0, n2-1]
        y2 = N[1, n2-1]

        L = np.sqrt((x2 - x1)**2 + (y2 - y1)**2)
        c = (x2 - x1) / L
        s = (y2 - y1) / L

        GDL = np.array([
            [2*n1 - 1],
            [2*n1],
            [2*n2 - 1],
            [2*n2]
        ])

        Ue = U[np.ix_(GDL.flatten())]
        De = (1 / L) * np.array([
            [-c, -s, c, s],
            [-s/L, c/L, s/L, -c/L]
        ]) @ Ue

        Epsi[i] = De[0]
        Fi[i] = E * A * De[0]
        Ti[i] = Fi[i] / A
    
    geraSaida('saida',Ft,Ut,Epsi,Fi,Ti)
    print('Ft = ', Ft)
    print('Ut = ', Ut)
    print('Epsi = ', Epsi)
    print('Fi = ', Fi)
    print('Ti = ', Ti)

    # Plot the structure with plota
    plota(N, Inc, Ut, Epsi, Fi, Ti)




